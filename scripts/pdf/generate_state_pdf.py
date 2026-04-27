"""
Generate state-specific Family Court Data Reports as PDF using the v2 HTML template.

Uses the same stat-computation logic as agent.py (state_stats, top_quotes,
children_impact) and renders via Jinja2 + WeasyPrint.

Usage:
    python generate_state_pdf.py --source=supabase MT
    python generate_state_pdf.py --source=supabase --all-30plus
    python generate_state_pdf.py --source=xlsx MT
"""
from __future__ import annotations
import os, sys, statistics
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta

import openpyxl
from jinja2 import Environment, FileSystemLoader
from playwright.sync_api import sync_playwright

BASE_DIR = Path(__file__).resolve().parent
MASTER_PATH = BASE_DIR / "outputs" / "SWM_MASTER_LATEST.xlsx"
TEMPLATE_DIR = BASE_DIR / "templates"
OUTPUT_DIR = BASE_DIR / "outputs" / "state_packets_pdf"
SHEET_NAME = "Sheet1"

COLS = {'state':1,'atty_fees':2,'gal_fees':3,'therapy_fees':4,'reunif_fees':5,
        'other_fees':6,'lost_wages':7,'asset_loss':8,'first_name':9,
        'permission':11,'quote':12,'case_status':13,'system':14,'duration':15,
        'custody':16,'pro_se':18,'legal_rep':19,'months_lost':21,'allegation':24,
        'county':30}

STATE_NAMES = {
    "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California","CO":"Colorado","CT":"Connecticut",
    "DE":"Delaware","FL":"Florida","GA":"Georgia","HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana",
    "IA":"Iowa","KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland","MA":"Massachusetts",
    "MI":"Michigan","MN":"Minnesota","MS":"Mississippi","MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada",
    "NH":"New Hampshire","NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina","ND":"North Dakota",
    "OH":"Ohio","OK":"Oklahoma","OR":"Oregon","PA":"Pennsylvania","RI":"Rhode Island","SC":"South Carolina",
    "SD":"South Dakota","TN":"Tennessee","TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington",
    "WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming","DC":"District of Columbia",
}

# Populate the state-token blocklist now that STATE_NAMES exists.
_STATE_TOKENS = {abbr.upper() for abbr in STATE_NAMES.keys()} | {
    name.upper() for name in STATE_NAMES.values()
}

STATE_COUNTIES = {
    "AL":67,"AK":30,"AZ":15,"AR":75,"CA":58,"CO":64,"CT":8,"DE":3,"FL":67,"GA":159,
    "HI":5,"ID":44,"IL":102,"IN":92,"IA":99,"KS":105,"KY":120,"LA":64,"ME":16,"MD":24,
    "MA":14,"MI":83,"MN":87,"MS":82,"MO":115,"MT":56,"NE":93,"NV":17,"NH":10,"NJ":21,
    "NM":33,"NY":62,"NC":100,"ND":53,"OH":88,"OK":77,"OR":36,"PA":67,"RI":5,"SC":46,
    "SD":66,"TN":95,"TX":254,"UT":29,"VT":14,"VA":133,"WA":39,"WV":55,"WI":72,"WY":23,"DC":1,
}

CUSTODY_NORM = {"Yes": "50/50 Joint", "No": "No Contact / Total Loss of Access"}

# Values that were entered into the County free-text field but are not
# counties — usually they are answers to an adjacent yes/no question, the
# state itself, or a country. They get filtered out of county aggregation
# so the report doesn't claim "Yes" or "Canada" is a Kansas county.
_COUNTY_JUNK = {
    'YES', 'NO', 'UNSURE', 'UNKNOWN', 'N/A', 'NA', 'NONE', 'NULL',
    'OTHER', 'INTERNATIONAL', 'INTERNATIONAL (LOCATION UNSPECIFIED)',
}
# Country names that show up in county fields when respondents are reporting
# from outside the US. Keep this list short and explicit.
_COUNTY_COUNTRY_JUNK = {
    'CANADA', 'MEXICO', 'UK', 'UNITED KINGDOM', 'AUSTRALIA',
    'GERMANY', 'FRANCE', 'IRELAND', 'NEW ZEALAND',
}

# _STATE_TOKENS — set of state abbreviations + full names used to reject
# 'CA' or 'California' typed where Los Angeles was meant. Populated above
# after STATE_NAMES is defined.

def safe_str(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s in ['None', 'nan', 'N/A'] else s


def normalize_county(v) -> str:
    """
    Whitespace-strip, title-case, and reject obvious non-county values.
    Returns "" for any value that should not appear in the county tally.
    """
    s = safe_str(v)
    if not s:
        return ""
    norm = " ".join(s.split())  # collapse internal whitespace
    upper = norm.upper()
    if upper in _COUNTY_JUNK or upper in _COUNTY_COUNTRY_JUNK:
        return ""
    if upper in _STATE_TOKENS:
        return ""
    # Title-case so "san diego" / "SAN DIEGO" / "San Diego " all collapse.
    # Preserve a trailing 'County' suffix if present without doubling it.
    title = norm.title()
    # Drop trailing " County" so "Johnson" and "Johnson County" merge.
    if title.endswith(" County"):
        title = title[: -len(" County")]
    return title


# System-Affected normalization: merge raw enum strings (e.g. "family_court")
# with their human-readable counterparts so the tally doesn't show two rows.
_SYSTEM_NORM = {
    'family_court': 'Family Court Only',
    'cps': 'CPS (Child Protective Services) Only',
    'both': 'Both Family Court and CPS',
}

def normalize_system(v) -> str:
    s = safe_str(v)
    if not s:
        return ""
    return _SYSTEM_NORM.get(s.strip().lower(), s)

def safe_float(v):
    try:
        if v is None or str(v).strip() in ['', 'None', 'N/A']:
            return None
        return float(str(v).replace('$', '').replace(',', '').strip())
    except Exception:
        return None

def normalize_custody(v):
    s = safe_str(v)
    return CUSTODY_NORM.get(s, s)

_CASE_STATUS_NORM = {
    'unknown': 'Not provided',
    'currently involved / "stuck"': 'Active case: stuck or delayed',
    'i am currently stuck in an active case (prolonged, delayed, or no resolution)': 'Active case: stuck or delayed',
    'active - progress': 'Active case: progress being made',
    'case closed (within the last 2 years)': 'Case closed: within last 2 years',
    'case closed (more than 2 years ago)': 'Case closed: more than 2 years ago',
    'experienced as a child': 'Experienced as a child',
    'i experienced this as a child in family court or cps': 'Experienced as a child',
    'i experienced this as a child in family court or child welfare / child protection agency': 'Experienced as a child',
    'not in court but still affected': 'Not in court but still affected',
}

def normalize_case_status(v):
    s = safe_str(v)
    if not s:
        return ''
    return _CASE_STATUS_NORM.get(s.lower(), s)

def load_survey(path):
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    return [list(r) for r in ws.iter_rows(values_only=True) if any(c for c in r)]

def build_states(rows):
    by_state = defaultdict(list)
    for row in rows:
        s = safe_str(row[1])
        if s and s not in {'None', 'nan', '', 'State of Occurrence', 'Unsure'} and len(s) <= 4:
            by_state[s.upper()].append(row)
    return by_state

def state_stats(rows):
    n = len(rows)
    if not n:
        return {}
    # Per-field sanity caps. Values above these are treated as data-entry
    # errors (e.g. someone typing a date into a months field, or adding an
    # extra zero to a dollar amount) and excluded from the stats.
    FIN_MAX = {
        'atty_fees':    5_000_000,
        'gal_fees':       500_000,
        'therapy_fees':   500_000,
        'reunif_fees':    500_000,
        'other_fees':     500_000,
        'lost_wages':   5_000_000,
        'asset_loss':  10_000_000,
        'months_lost':        240,   # 20 years
    }
    def fin(k):
        cap = FIN_MAX.get(k)
        vals = []
        for r in rows:
            if COLS[k] >= len(r):
                continue
            value = safe_float(r[COLS[k]])
            if value is None or value < 0:
                continue
            if cap is not None and value > cap:
                continue
            vals.append(value)
        if not vals:
            return {'mean': None, 'median': None, 'sum': None, 'n': 0}
        return {
            'mean': round(sum(vals) / len(vals)),
            'median': round(statistics.median(vals)),
            'sum': round(sum(vals)),
            'n': len(vals),
        }
    def top(k, t=8):
        if k == 'custody':
            c = Counter(normalize_custody(r[COLS[k]]) for r in rows if normalize_custody(r[COLS[k]]))
        elif k == 'county':
            c = Counter(normalize_county(r[COLS[k]]) for r in rows if normalize_county(r[COLS[k]]))
        elif k == 'system':
            c = Counter(normalize_system(r[COLS[k]]) for r in rows if normalize_system(r[COLS[k]]))
        elif k == 'case_status':
            c = Counter(normalize_case_status(r[COLS[k]]) for r in rows if normalize_case_status(r[COLS[k]]))
        else:
            c = Counter(safe_str(r[COLS[k]]) for r in rows if safe_str(r[COLS[k]]))
        return c.most_common(t)
    prose = sum(1 for r in rows if 'yes' in safe_str(r[COLS['pro_se']]).lower() or 'pro se' in safe_str(r[COLS['pro_se']]).lower())
    ran = sum(1 for r in rows if 'ran out' in safe_str(r[COLS['legal_rep']]).lower())
    # Count all unique normalized counties for the cover-page stat. The
    # top-N list is capped, so deriving counties_represented from len(top)
    # would always max out at the cap.
    unique_counties = {
        normalize_county(r[COLS['county']])
        for r in rows
        if normalize_county(r[COLS['county']])
    }
    return {
        'total': n, 'pro_se_pct': round(prose / n * 100, 1), 'ran_out_pct': round(ran / n * 100, 1),
        'financial': {k: fin(k) for k in ['atty_fees', 'gal_fees', 'therapy_fees', 'reunif_fees', 'other_fees', 'lost_wages', 'asset_loss']},
        'months_lost': fin('months_lost'),
        'case_status': top('case_status'), 'system': top('system'), 'duration': top('duration'),
        'custody': top('custody'), 'allegations': top('allegation', 10), 'counties': top('county', 10),
        'counties_unique': len(unique_counties),
    }

def children_impact(rows):
    total_children = 0
    submissions_with_children = 0
    for r in rows:
        if 17 >= len(r):
            continue
        raw = safe_str(r[17])
        if not raw:
            continue
        try:
            n = int(float(raw))
        except Exception:
            continue
        if n < 0 or n > 20:
            continue
        total_children += n
        submissions_with_children += 1
    avg = round(total_children / submissions_with_children, 1) if submissions_with_children else None
    return {
        'total_children': total_children,
        'submissions_with_children': submissions_with_children,
        'avg_children': avg,
    }

def score_quote(q, p):
    if not q or len(q) < 20:
        return 0
    pl = p.lower()
    if 'do not share' in pl or 'data purposes' in pl:
        return 0
    perm = 3 if 'share away' in pl else 2 if 'first name' in pl else 1 if 'anonymous' in pl else 0
    if not perm:
        return 0
    power = ['took my', 'stolen', 'ripped', 'no trial', 'no due process', 'false', 'corrupt', 'destroyed',
             'children', 'kids', 'silence', 'fraud', 'abuse', 'pay to play', 'gal', 'judge', 'cps',
             'lost everything', 'no contact', 'trauma', 'rights', 'attorney']
    words = sum(2 for w in power if w in q.lower())
    l = len(q)
    length = 5 if 30 <= l <= 150 else 3 if l <= 300 else 1
    return perm + words + length

QUOTE_MAX_CHARS = 280  # keeps voice cards a uniform size so they don't bleed off the page

def _trim_quote(q):
    q = " ".join(q.split())  # collapse whitespace
    if len(q) <= QUOTE_MAX_CHARS:
        return q
    cut = q[:QUOTE_MAX_CHARS].rsplit(' ', 1)[0].rstrip(' ,;:-')
    return cut + " […]"

def top_quotes(rows, n=None):
    scored = []
    for r in rows:
        q = safe_str(r[COLS['quote']])
        full_q = " ".join(q.split())
        p = safe_str(r[COLS['permission']])
        fn = safe_str(r[COLS['first_name']])
        if 'do not share' in p.lower() or 'data purposes' in p.lower():
            continue
        s = score_quote(q, p)
        if s > 0 and q:
            pl = p.lower()
            attr = fn if ('share away' in pl or 'first name' in pl) and fn not in ['None', '', 'nan'] else "Anonymous Parent"
            scored.append({
                'quote': _trim_quote(q),
                'full_quote': full_q,
                'score': s,
                'attribution': attr,
                'permission': p,
            })
    scored.sort(key=lambda x: x['score'], reverse=True)
    return scored if n is None else scored[:n]


FULL_QUOTE_CHARS_PER_LINE = 92
FULL_QUOTE_LINE_H = 15
FULL_QUOTE_CARD_VPAD = 34
FULL_QUOTE_CITE_LINE = 20
FULL_QUOTE_GAP = 10
FULL_QUOTE_PAGE_BUDGET = 730
FULL_QUOTE_CHUNK_CHARS = 1700


def _split_quote_text(text, max_chars=FULL_QUOTE_CHUNK_CHARS):
    """Split unusually long quotes across appendix cards without truncating."""
    text = " ".join(str(text or "").split())
    if len(text) <= max_chars:
        return [text] if text else []
    chunks = []
    remaining = text
    while len(remaining) > max_chars:
        cut = remaining[:max_chars].rsplit(" ", 1)[0].rstrip()
        if len(cut) < max_chars * 0.6:
            cut = remaining[:max_chars].rstrip()
        chunks.append(cut)
        remaining = remaining[len(cut):].strip()
    if remaining:
        chunks.append(remaining)
    return chunks


def _full_quote_card_height(text):
    chars = len(text or "")
    lines = max(1, -(-chars // FULL_QUOTE_CHARS_PER_LINE))
    return lines * FULL_QUOTE_LINE_H + FULL_QUOTE_CITE_LINE + FULL_QUOTE_CARD_VPAD


def full_quote_pages(quotes):
    """Pack every approved/shareable full quote into appendix pages."""
    cards = []
    for q in quotes:
        chunks = _split_quote_text(q.get('full_quote') or q.get('quote') or "")
        for idx, chunk in enumerate(chunks):
            attribution = q['attribution']
            if len(chunks) > 1:
                attribution = f"{attribution} ({idx + 1}/{len(chunks)})"
            cards.append({
                'quote': chunk,
                'attribution': attribution,
            })

    pages = []
    page = []
    used = 0
    for card in cards:
        h = _full_quote_card_height(card['quote'])
        if page and used + h + FULL_QUOTE_GAP > FULL_QUOTE_PAGE_BUDGET:
            pages.append({'cards': page})
            page = []
            used = 0
        page.append(card)
        used += h + FULL_QUOTE_GAP
    if page:
        pages.append({'cards': page})
    return pages


def fmt_dollar(v):
    if v is None:
        return "—"
    return f"${v:,}"

def fmt_num(v):
    if v is None:
        return "N/A"
    return f"{v:,}" if isinstance(v, int) else str(v)


def _bar_items(top_list, total):
    """Convert a list of (label, count) tuples into template-ready dicts with bar_pct."""
    if not top_list:
        return []
    max_count = top_list[0][1] if top_list else 1
    items = []
    for label, count in top_list:
        pct = round(count / total * 100) if total else 0
        bar_pct = round(count / max_count * 100) if max_count else 0
        items.append({'label': label, 'count': count, 'pct': pct, 'bar_pct': bar_pct})
    return items


def _check_outlier(vals):
    """Return True if the max value is more than 3 standard deviations above the mean."""
    if len(vals) < 3:
        return False
    mean = sum(vals) / len(vals)
    sd = statistics.stdev(vals)
    return sd > 0 and (max(vals) - mean) / sd > 3


def _get_asset_loss_vals(rows):
    vals = []
    for r in rows:
        if COLS['asset_loss'] >= len(r):
            continue
        v = safe_float(r[COLS['asset_loss']])
        if v is not None:
            vals.append(v)
    return vals


def build_template_context(state_abbr, rows, court_actors=None):
    """Build the full dict of template variables for a state."""
    court_actors = court_actors or []
    d = state_stats(rows)
    n = d['total']
    kids = children_impact(rows)
    f = d['financial']
    ml = d['months_lost']
    state_name = STATE_NAMES.get(state_abbr, state_abbr)
    now = datetime.now(timezone(timedelta(hours=-5)))

    counties_represented = d.get('counties_unique', len(d['counties']))

    # Expense table rows
    asset_vals = _get_asset_loss_vals(rows)
    has_asset_outlier = _check_outlier(asset_vals)

    expense_categories = [
        ("Attorney Fees", 'atty_fees', None),
        ("Lost Wages", 'lost_wages', None),
        ("Other Court Fees", 'other_fees', None),
        ("Therapy / Evaluations", 'therapy_fees', None),
        ("Reunification Services", 'reunif_fees', None),
        ("GAL Fees", 'gal_fees', None),
        ("Asset Loss", 'asset_loss', "1" if has_asset_outlier else None),
    ]

    expense_rows = []
    median_sum = 0
    for label, key, footnote in expense_categories:
        fi = f[key]
        if key == 'asset_loss' and has_asset_outlier:
            row = {
                'label': label,
                'n': fi['n'],
                'median': fmt_dollar(fi['median']),
                'mean': "—",
                'total': "—",
                'footnote': footnote,
            }
        else:
            row = {
                'label': label,
                'n': fi['n'],
                'median': fmt_dollar(fi['median']),
                'mean': fmt_dollar(fi['mean']),
                'total': fmt_dollar(fi['sum']),
                'footnote': footnote,
            }
        expense_rows.append(row)
        if fi['median'] is not None:
            median_sum += fi['median']

    asset_footnote_text = ""
    if has_asset_outlier:
        asset_footnote_text = (
            f"One respondent reported an extreme outlier that skews the mean beyond usefulness. "
            f"We report the median ({fmt_dollar(f['asset_loss']['median'])}) and suppress the mean and total, "
            f"pending data-quality review. Lost Wages also contains wide variance; median is the honest headline."
        )

    # Months lost bars — scale relative to total
    months_total = ml['sum'] if ml['sum'] else 0
    months_median = ml['median'] if ml['median'] else 0
    months_mean = ml['mean'] if ml['mean'] else 0
    max_months = max(months_total, 1)

    def yrs_label(months):
        if months is None or months == 0:
            return "0"
        y = round(months / 12, 1)
        if y == int(y):
            return f"{int(y)} {'yr' if int(y) == 1 else 'yrs'}"
        return f"{y} yrs"

    # Allegation quote — pick first quote that mentions due process/evidence/gag/unconstitutional.
    # Reserve this BEFORE building voice_pages so the same quote doesn't get shown twice.
    allegation_quote = None
    allegation_keywords = ['due process', 'evidence', 'gag order', 'unconstitutional', 'ex parte', 'denied']
    all_quotes = top_quotes(rows)
    for q in all_quotes:
        ql = (q.get('full_quote') or q['quote']).lower()
        if any(kw in ql for kw in allegation_keywords):
            allegation_quote = {
                'quote': q['quote'],
                'attribution': f"{state_name} parent · {q['attribution'].lower() if q['attribution'] != 'Anonymous Parent' else 'anonymous'}",
            }
            break
    total_publishable = len(all_quotes)
    # Drop the allegation quote from the main pool so it doesn't reappear as
    # a pullquote or card on the voices pages.
    reserved_text = allegation_quote['quote'] if allegation_quote else None
    voice_pool = [q for q in all_quotes if q['quote'] != reserved_text]

    # Quotes — dynamically pack as many voice cards as fit on each page.
    # Empirical budgets in pixels for the 2-column grid area after the
    # pullquote and section headings have been laid out. Tuned conservatively
    # because pullquotes with 3-4 lines of text eat more vertical space than
    # the naive estimate (and we'd rather waste a few px than bleed a card
    # into the page-foot stamp).
    #   - FIRST voices page: pullquote + allegation-quote block + 2 subsection headings
    #   - SUBSEQUENT voices pages: pullquote + cards
    FIRST_PAGE_GRID_PX = 420
    PAGE_GRID_PX       = 620
    CHARS_PER_LINE     = 56   # slightly conservative
    LINE_H             = 18   # 12px × 1.45 line-height, rounded up
    CARD_VPAD          = 24   # padding + attribution margin
    CITE_LINE          = 20   # attribution line + margin
    CARD_SAFETY        = 10   # extra slack per card for borders/rendering
    GRID_GAP           = 8

    def _card_height(q):
        chars = len(q['quote'])
        lines = max(1, -(-chars // CHARS_PER_LINE))  # ceil division
        return lines * LINE_H + CITE_LINE + CARD_VPAD + CARD_SAFETY

    voice_pages = []
    quotes_shown = 1 if allegation_quote else 0  # allegation quote on first voices page
    remaining = list(voice_pool)
    while remaining and len(voice_pages) < 5:
        budget = FIRST_PAGE_GRID_PX if not voice_pages else PAGE_GRID_PX
        page = {'pullquote': remaining.pop(0) if remaining else None, 'cards': []}
        if page['pullquote']:
            quotes_shown += 1
        used = 0
        while len(remaining) >= 1:
            left = remaining[0]
            right = remaining[1] if len(remaining) > 1 else None
            row_h = max(_card_height(left), _card_height(right) if right else 0) + GRID_GAP
            if used + row_h > budget:
                break
            page['cards'].append(remaining.pop(0))
            quotes_shown += 1
            if right:
                page['cards'].append(remaining.pop(0))
                quotes_shown += 1
            used += row_h
        # Safety: always publish at least one card per page so we don't get
        # stuck in an infinite loop on an unusually tall quote.
        if not page['cards'] and remaining:
            page['cards'].append(remaining.pop(0))
            quotes_shown += 1
        voice_pages.append(page)

    num_voice_pages = len(voice_pages)
    full_comment_pages = full_quote_pages(all_quotes)
    num_full_comment_pages = len(full_comment_pages)
    actor_page_count = 1 if court_actors else 0
    voices_start_page = 5 + actor_page_count
    full_comments_start_page = voices_start_page + num_voice_pages
    methodology_page_number = full_comments_start_page + num_full_comment_pages
    total_pages = methodology_page_number  # cover + glance + financial + who-geo-alleg + actors + voices + full comments + methodology

    # State number (alphabetical order of state abbreviation)
    sorted_states = sorted(STATE_NAMES.keys())
    state_number = sorted_states.index(state_abbr) + 1 if state_abbr in sorted_states else 0

    return {
        'state_name': state_name,
        'state_abbr': state_abbr,
        'state_number': state_number,
        'total': n,
        'report_date': now.strftime('%B %d, %Y').replace(' 0', ' '),
        'report_month_year': now.strftime('%B %Y'),
        'total_pages': f"{total_pages:02d}",
        'actor_page_number': 5,
        'voices_start_page': voices_start_page,
        'full_comments_start_page': full_comments_start_page,
        'methodology_page_number': methodology_page_number,

        # Children
        'children_total': kids['total_children'],
        'children_reporting': kids['submissions_with_children'],
        'avg_children': kids['avg_children'] if kids['avg_children'] is not None else 'N/A',

        # Counties
        'counties_represented': counties_represented,
        'counties_in_state': STATE_COUNTIES.get(state_abbr, '?'),

        # Headline stats
        'atty_fees_median': fmt_dollar(f['atty_fees']['median']),
        'pro_se_pct': d['pro_se_pct'],
        'ran_out_pct': d['ran_out_pct'],

        # Months lost
        'months_lost_median': months_median,
        'months_lost_mean': months_mean,
        'months_lost_total': f"{months_total:,}",
        'months_lost_median_bar': round(months_median / max_months * 100) if max_months else 0,
        'months_lost_mean_bar': round(months_mean / max_months * 100) if max_months else 0,
        'months_lost_median_yrs': yrs_label(months_median),
        'months_lost_mean_yrs': yrs_label(months_mean),
        'months_lost_total_yrs': yrs_label(months_total),

        # Financial table
        'expense_rows': expense_rows,
        'median_family_impact': fmt_dollar(median_sum),
        'has_asset_footnote': has_asset_outlier,
        'asset_footnote_text': asset_footnote_text,

        # Bar chart sections
        'case_status': _bar_items(d['case_status'][:5], n),
        'system': _bar_items(d['system'][:5], n),
        'duration': _bar_items(d['duration'][:5], n),
        'custody': _bar_items(d['custody'][:5], n),
        'counties': _bar_items(d['counties'][:6], n),
        'allegations': _bar_items(d['allegations'][:5], n),

        # Court actors
        'court_actors': court_actors[:18],
        'court_actor_count': len(court_actors),
        'court_actor_threshold': 5,

        # Quotes
        'allegation_quote': allegation_quote,
        'voice_pages': voice_pages,
        'total_publishable_quotes': total_publishable,
        'quotes_shown': quotes_shown,
        'full_comment_pages': full_comment_pages,
        'full_comments_total': total_publishable,
    }


def generate_pdf(state_abbr, rows, output_path=None, browser=None, court_actors=None):
    """Render the Jinja2 template and convert to PDF via Playwright."""
    ctx = build_template_context(state_abbr, rows, court_actors=court_actors)

    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template('state_report.html.j2')
    html_str = template.render(**ctx)

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"{state_abbr}.pdf"

    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix='.html', dir=str(TEMPLATE_DIR), delete=False)
    tmp.write(html_str.encode('utf-8'))
    tmp.close()

    own_browser = browser is None
    pw = None
    used_renderer = "Playwright"
    try:
        try:
            if os.environ.get("SWM_PDF_RENDERER", "").lower() == "weasyprint":
                raise RuntimeError("SWM_PDF_RENDERER=weasyprint")
            else:
                if own_browser:
                    pw = sync_playwright().start()
                    browser = pw.chromium.launch()
                page = browser.new_page()
                page.goto(f"file://{tmp.name}", wait_until="networkidle")
                page.pdf(
                    path=str(output_path),
                    format="Letter",
                    print_background=True,
                    margin={"top": "0in", "bottom": "0in", "left": "0in", "right": "0in"},
                )
                page.close()
        except Exception as playwright_err:
            # Local Codex/macOS sandboxes can block Chromium's Mach port
            # registration. GitHub Actions should use Playwright; this fallback
            # keeps local external-drive runs usable when WeasyPrint is present.
            if browser is not None and not own_browser:
                raise
            print(f"  Playwright failed, trying WeasyPrint fallback: {playwright_err}")
            try:
                from weasyprint import HTML
            except Exception as import_err:
                raise playwright_err from import_err
            HTML(filename=tmp.name, base_url=str(TEMPLATE_DIR)).write_pdf(str(output_path))
            used_renderer = "WeasyPrint"
    finally:
        Path(tmp.name).unlink(missing_ok=True)
        if own_browser and browser is not None:
            browser.close()
        if own_browser and pw is not None:
            pw.stop()

    print(f"  PDF written: {output_path} ({output_path.stat().st_size / 1024:.0f} KB, {used_renderer})")
    return output_path


def main():
    args = [a for a in sys.argv[1:]]
    use_supabase = True
    only_30plus = False
    cleaned: list[str] = []
    for a in args:
        if a == "--source=supabase":
            use_supabase = True
        elif a == "--source=xlsx":
            use_supabase = False
        elif a == "--all-30plus":
            only_30plus = True
        else:
            cleaned.append(a)
    args = cleaned

    if not args and not only_30plus:
        print("Usage: python generate_state_pdf.py [--source=supabase|xlsx] "
              "[--all-30plus] <STATE> [STATE ...]")
        sys.exit(1)

    if use_supabase:
        from lib_supabase_rows import load_public_court_actors_from_supabase, load_rows_from_supabase
        print("Loading rows from Supabase (survey_submissions + legacy_submissions)")
        rows = load_rows_from_supabase()
        public_actors_by_state = load_public_court_actors_from_supabase()
    else:
        print(f"Loading master workbook: {MASTER_PATH}")
        rows = load_survey(MASTER_PATH)
        public_actors_by_state = {}

    by_state = build_states(rows)
    print(f"  {len(rows):,} total rows, {len(by_state)} states")

    # Resolve target state list
    if only_30plus:
        states = sorted(s for s, rs in by_state.items() if len(rs) >= 30)
        print(f"  --all-30plus: {len(states)} states qualify")
    else:
        states = [s.strip().upper() for s in args]

    generated = 0
    for state in states:
        if state not in by_state:
            print(f"  {state}: not found in data, skipping")
            continue
        sr = by_state[state]
        if len(sr) < 30:
            print(f"  {state}: {len(sr)} submissions, below 30-family threshold; skipping")
            continue
        print(f"\n  Generating {STATE_NAMES.get(state, state)} ({state}) — {len(sr)} submissions")
        generate_pdf(state, sr, court_actors=public_actors_by_state.get(state, []))
        generated += 1

    print(f"\nDone. Generated {generated} PDF(s).")


if __name__ == "__main__":
    main()
